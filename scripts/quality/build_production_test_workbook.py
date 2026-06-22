#!/usr/bin/env python3
"""Nâng cấp test catalogue AVM thành workbook production release-gate.

Workbook đầu vào là bản audit hoặc catalogue cũ có sheet ``Test Case Sản xuất``.
Workbook đầu ra tách test case, happy path, failure path, automation mapping,
execution evidence và RTM để một test không còn nhét nhiều kịch bản vào một ô.
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_TARGET = PROJECT_ROOT / "docs" / "testing" / "AVM_Production_Test_Cases.xlsx"

LEGACY_HEADERS = (
    "STT",
    "Mã test",
    "Cấp kiểm thử",
    "Loại kiểm thử",
    "Khu vực",
    "Kịch bản",
    "Điều kiện trước",
    "Bước thực hiện",
    "Kết quả mong đợi",
    "Tự động hóa",
    "Ưu tiên",
    "Trạng thái",
    "Liên kết kiểm chứng",
)

FIXED_CASES = {
    "ML-LIN-017": {
        "Khu vực": "ML lineage",
        "Kịch bản": "Giữ nguyên evidence_tier từ PostgreSQL khi train",
        "Điều kiện trước": "properties có evidence_tier E2-E5",
        "Bước thực hiện": "Gọi load_data_from_db; đối chiếu phân bố tier với truy vấn PostgreSQL",
        "Kết quả mong đợi": "Phân bố tier khớp DB; pipeline không tự suy luận lại khi cột đã tồn tại",
        "Tự động hóa": "pytest",
        "Ưu tiên": "P0",
        "Trạng thái": "Automated",
        "Liên kết kiểm chứng": "tests/unit/test_pipeline_load_data.py",
    },
    "ML-LIN-018": {
        "Khu vực": "ML benchmark",
        "Kịch bản": "So sánh model trên cùng holdout và checksum",
        "Điều kiện trước": "Run mới có split manifest và dataset checksum",
        "Bước thực hiện": "Retrain; đối chiếu exact holdout/checksum trước khi so metric",
        "Kết quả mong đợi": "Chỉ cho phép so trực tiếp khi cùng holdout hoặc có mapping sai khác rõ ràng",
        "Tự động hóa": "pytest + PostgreSQL",
        "Ưu tiên": "P0",
        "Trạng thái": "Ready",
        "Liên kết kiểm chứng": "src/ml/pipeline.py; tests/production/test_ml_release_gate.py",
    },
    "CLEAN-004": {
        "Khu vực": "Documentation",
        "Kịch bản": "Runbook phản ánh runtime PostgreSQL-only",
        "Điều kiện trước": "Tài liệu và cấu hình production đã cập nhật",
        "Bước thực hiện": "Đối chiếu README, production runbook, Docker, Alembic và testing docs",
        "Kết quả mong đợi": "Không hướng dẫn SQLite runtime; có PostgreSQL, Alembic, CI, Docker và OAuth env",
        "Tự động hóa": "pytest contract + review",
        "Ưu tiên": "P1",
        "Trạng thái": "Automated",
        "Liên kết kiểm chứng": "tests/unit/test_postgresql_only_contract.py; docs/runbooks/SPEC-PRODUCTION.md",
    },
}

DUPLICATE_IDS = {
    ("UI-UX-001", 1): "UI-AUDIT-001",
    ("UI-UX-001", 2): "UI-LOGIN-001",
    ("UI-UX-002", 1): "UI-PRED-001",
    ("UI-UX-002", 2): "UI-PRED-002",
    ("UI-UX-003", 1): "UI-OAUTH-001",
    ("UI-UX-003", 2): "UI-RBAC-001",
    ("CI-001", 1): "CI-GATE-001",
    ("CI-001", 2): "CI-BE-001",
    ("CI-002", 1): "CI-SONAR-001",
    ("CI-002", 2): "CI-FE-001",
    ("CI-003", 1): "CI-FE-002",
    ("CI-003", 2): "CI-DOCKER-001",
    ("CI-004", 1): "CI-SEC-001",
    ("CI-004", 2): "CI-SONAR-002",
    ("CI-005", 1): "CI-CATALOG-001",
    ("CI-005", 2): "CI-COVERAGE-001",
    ("DOCKER-001", 1): "DOCKER-STACK-001",
    ("DOCKER-001", 2): "DOCKER-BE-001",
    ("DOCKER-002", 1): "DOCKER-SEC-001",
    ("DOCKER-002", 2): "DOCKER-MIG-001",
    ("SEC-001", 1): "SEC-HEADER-001",
    ("SEC-001", 2): "SEC-SECRET-001",
}

AUTOMATED_REFERENCES = {
    "AUTH-WB-003": "tests/unit/test_google_oauth.py; tests/unit/test_health_endpoint.py",
    "ML-LIN-017": "tests/unit/test_pipeline_load_data.py",
    "ML-LIN-018": "tests/production/test_ml_release_gate.py",
    "CI-CATALOG-001": "scripts/quality/validate_test_catalog.py",
    "DOCKER-BE-001": "tests/production/test_api_release_gate.py",
    "SEC-SECRET-001": "tests/unit/test_postgresql_only_contract.py",
    "CLEAN-004": "tests/unit/test_postgresql_only_contract.py",
}

THIN = Side(style="thin", color="D9E2EA")
HEADER_FILL = PatternFill("solid", fgColor="123047")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCEBF2")
ACCENT_FILL = PatternFill("solid", fgColor="D7F0E8")
WARNING_FILL = PatternFill("solid", fgColor="FFF0C2")
DANGER_FILL = PatternFill("solid", fgColor="F9D6D5")
WHITE_FONT = Font(color="FFFFFF", bold=True)


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _load_legacy_cases(source: Path) -> list[dict[str, str]]:
    workbook = load_workbook(source, data_only=False)
    if "Test Case Sản xuất" not in workbook.sheetnames:
        raise ValueError(f"{source} không có sheet 'Test Case Sản xuất'")
    sheet = workbook["Test Case Sản xuất"]
    headers = tuple(_clean(sheet.cell(1, column).value) for column in range(1, 14))
    if headers != LEGACY_HEADERS:
        raise ValueError(f"Header catalogue cũ không hợp lệ: {headers}")

    cases: list[dict[str, str]] = []
    occurrences: defaultdict[str, int] = defaultdict(int)
    for row in range(2, sheet.max_row + 1):
        raw = {
            header: _clean(sheet.cell(row, index).value)
            for index, header in enumerate(LEGACY_HEADERS, start=1)
        }
        old_id = raw["Mã test"]
        if not old_id:
            continue
        occurrences[old_id] += 1
        raw.update(FIXED_CASES.get(old_id, {}))
        raw["Mã test"] = DUPLICATE_IDS.get((old_id, occurrences[old_id]), old_id)
        cases.append(raw)
    workbook.close()

    ids = [case["Mã test"] for case in cases]
    duplicates = [test_id for test_id, count in Counter(ids).items() if count > 1]
    if duplicates:
        raise ValueError(f"Mã test vẫn bị trùng sau chuẩn hóa: {duplicates}")
    return cases


def _requirement_id(test_id: str) -> str:
    prefix = test_id.split("-", 1)[0]
    return f"REQ-{prefix}-{test_id.rsplit('-', 1)[-1]}"


def _risk_category(case: dict[str, str]) -> str:
    text = " ".join(case.values()).lower()
    if any(token in text for token in ("oauth", "jwt", "secret", "security", "rbac", "auth")):
        return "Security & Access"
    if any(token in text for token in ("mape", "mae", "model", "holdout", "retrain", "lineage", "ml")):
        return "Model Risk"
    if any(token in text for token in ("postgres", "alembic", "database", "schema", "sql")):
        return "Data Integrity"
    if any(token in text for token in ("docker", "ci", "github", "sonar", "deploy")):
        return "Release & Operations"
    if any(token in text for token in ("ui", "ux", "responsive", "playwright")):
        return "User Experience"
    if any(token in text for token in ("latency", "performance", "load", "200ms")):
        return "Performance"
    return "Functional Correctness"


def _normalize_priority(case: dict[str, str]) -> str:
    priority = case["Ưu tiên"].upper()
    if priority not in {"P0", "P1", "P2", "P3"}:
        return "P1"
    return priority


def _automation_status(case: dict[str, str]) -> tuple[str, str]:
    test_id = case["Mã test"]
    reference = AUTOMATED_REFERENCES.get(test_id, case["Liên kết kiểm chứng"])
    status = case["Trạng thái"].lower()
    reference_lower = reference.lower()
    has_executable_ref = any(
        token in reference_lower
        for token in ("tests/", ".github/workflows/", "scripts/quality/", "npm run", "pytest")
    )
    if status == "automated" and has_executable_ref:
        return "Automated", reference
    if has_executable_ref or status in {"automated", "ready"}:
        return "Partial", reference
    return "Planned", reference


def _failure_templates(case: dict[str, str]) -> list[dict[str, str]]:
    text = " ".join(case.values()).lower()
    subject = case["Kịch bản"]
    expected = case["Kết quả mong đợi"]

    if any(token in text for token in ("login", "oauth", "jwt", "refresh", "auth", "rbac")):
        return [
            {
                "title": f"{subject} - credential/token không hợp lệ",
                "trigger": "Token hết hạn, state replay, sai credential hoặc account bị khóa",
                "steps": f"Chuẩn bị dữ liệu xác thực không hợp lệ; thực hiện: {case['Bước thực hiện']}; thử replay request",
                "expected": "Fail closed với 401/403/429 phù hợp; không tạo session/token; log không lộ secret",
                "recovery": "Thu hồi token liên quan; kiểm tra audit log; cấp lại phiên qua luồng đăng nhập chuẩn sau lockout",
            },
            {
                "title": f"{subject} - thiếu cấu hình hoặc upstream OAuth lỗi",
                "trigger": "Thiếu env, callback sai, timeout hoặc provider trả lỗi",
                "steps": "Tắt/mask env bắt buộc hoặc mock provider timeout; gọi lại endpoint; kiểm tra response và log",
                "expected": "503/4xx có thông báo an toàn; không redirect vòng lặp; không ghi raw secret/token",
                "recovery": "Khôi phục env từ secret store; xác minh redirect allowlist; retry có kiểm soát sau health check",
            },
        ]
    if any(token in text for token in ("postgres", "alembic", "database", "schema", "lineage", "history")):
        return [
            {
                "title": f"{subject} - dữ liệu vi phạm constraint",
                "trigger": "Thiếu FK, trùng unique key, JSON sai cấu trúc hoặc giá trị ngoài miền",
                "steps": f"Gửi fixture vi phạm constraint vào transaction; thực hiện: {case['Bước thực hiện']}",
                "expected": "Transaction rollback toàn bộ; dữ liệu chính không bẩn; lỗi có mã rõ và row lỗi được quarantine khi phù hợp",
                "recovery": "Sửa fixture/nguồn; đối chiếu constraint và migration; chạy lại trong transaction mới sau data-quality gate",
            },
            {
                "title": f"{subject} - PostgreSQL/migration không sẵn sàng",
                "trigger": "DB mất kết nối, migration chưa ở head hoặc pool cạn",
                "steps": "Mô phỏng kết nối thất bại hoặc revision lệch; gọi health/readiness và thao tác liên quan",
                "expected": "Readiness fail rõ; app không tự tạo schema và không fallback SQLite; không có partial commit",
                "recovery": "Khôi phục PostgreSQL; chạy Alembic upgrade head; kiểm tra pool/lock rồi mới mở traffic",
            },
        ]
    if any(token in text for token in ("mape", "mae", "model", "holdout", "retrain", "drift", "ml")):
        return [
            {
                "title": f"{subject} - data/model quality gate không đạt",
                "trigger": "Thiếu evidence tier, checksum lệch, leakage, drift cao hoặc metric candidate kém production",
                "steps": f"Dùng fixture vi phạm quality policy; thực hiện: {case['Bước thực hiện']}; yêu cầu promote candidate",
                "expected": "Training/promotion bị chặn; active model không đổi; lưu metric, checksum và lý do reject",
                "recovery": "Chuẩn hóa dataset/split; loại leakage; retrain cùng holdout; review metric trước khi promote lại",
            },
            {
                "title": f"{subject} - artifact thiếu hoặc không khớp registry",
                "trigger": "Model file thiếu, hash sai, metadata version khác DB hoặc load model lỗi",
                "steps": "Đổi fixture artifact/hash; khởi tạo model loader và gọi prediction smoke",
                "expected": "Không phục vụ artifact không xác minh; health/degraded state rõ; không tự chọn candidate mới nhất",
                "recovery": "Khôi phục artifact active từ registry/backup; xác minh SHA-256; clear cache và smoke test trước mở traffic",
            },
        ]
    if any(token in text for token in ("docker", "github", "ci", "sonar", "pipeline", "deploy")):
        return [
            {
                "title": f"{subject} - build/test/security gate thất bại",
                "trigger": "Một test fail, vulnerability vượt ngưỡng, workbook sai contract hoặc image build lỗi",
                "steps": f"Chèn lỗi kiểm soát; trigger pipeline; thực hiện gate: {case['Bước thực hiện']}",
                "expected": "Workflow đỏ; job sau không deploy; JUnit/log/artifact được giữ để truy nguyên",
                "recovery": "Tạo defect từ evidence; sửa nguyên nhân; chạy lại đúng commit SHA; không bypass required check",
            },
            {
                "title": f"{subject} - service/secret dependency không sẵn sàng",
                "trigger": "PostgreSQL unhealthy, migration fail hoặc thiếu GitHub Secret bắt buộc",
                "steps": "Khởi động stack/workflow với dependency lỗi; quan sát health, exit code và cleanup",
                "expected": "Fail fast có thông báo; backend không serve trước migration; container/job được cleanup",
                "recovery": "Khôi phục dependency/secret; rotate nếu lộ; chạy compose config và smoke test trước retry",
            },
        ]
    if any(token in text for token in ("ui", "ux", "responsive", "playwright", "frontend")):
        return [
            {
                "title": f"{subject} - error/empty/loading state",
                "trigger": "API trả 4xx/5xx, dữ liệu rỗng hoặc response chậm",
                "steps": f"Mock từng trạng thái; thực hiện: {case['Bước thực hiện']}; chụp desktop và mobile",
                "expected": "Không blank/overlap/layout shift; có loading, empty và error action rõ; quyền user/admin đúng",
                "recovery": "Cho phép retry/cancel an toàn; giữ input người dùng; ghi client error correlation ID",
            },
            {
                "title": f"{subject} - dữ liệu dài và viewport biên",
                "trigger": "Text dài, số VND lớn, viewport 320px hoặc zoom 200%",
                "steps": "Dùng fixture tiếng Việt dài và giá trị biên; chạy visual/a11y smoke trên viewport nhỏ/lớn",
                "expected": "Text wrap đúng; control không tràn; focus/keyboard hoạt động; không che nội dung quan trọng",
                "recovery": "Điều chỉnh responsive constraints; cập nhật snapshot có review; chạy lại toàn bộ route cùng role",
            },
        ]
    if any(token in text for token in ("latency", "performance", "load", "200ms", "cache")):
        return [
            {
                "title": f"{subject} - vượt SLO latency",
                "trigger": "p95 >= 200ms, timeout hoặc error rate vượt ngưỡng",
                "steps": f"Warm-up rồi chạy tải có kiểm soát; thực hiện: {case['Bước thực hiện']}; thu p50/p95/p99",
                "expected": "Gate fail khi vượt SLO; evidence có latency histogram, error rate và commit SHA",
                "recovery": "Profile query/render; kiểm tra index/cache/pool; tối ưu và chạy lại cùng workload",
            },
            {
                "title": f"{subject} - tải đồng thời và backpressure",
                "trigger": "Concurrency tăng, client hủy request hoặc pool đạt giới hạn",
                "steps": "Tăng concurrency theo bậc; hủy một phần request; quan sát resource và recovery",
                "expected": "Không leak connection/memory; timeout/rate limit có kiểm soát; hệ thống hồi phục sau tải",
                "recovery": "Giảm traffic; điều chỉnh pool/timeout; xác minh cancellation và chạy soak test ngắn",
            },
        ]

    return [
        {
            "title": f"{subject} - input thiếu hoặc ngoài miền",
            "trigger": "Thiếu trường bắt buộc, sai type, boundary hoặc dữ liệu không nhất quán",
            "steps": f"Biến đổi fixture hợp lệ thành invalid/boundary; thực hiện: {case['Bước thực hiện']}",
            "expected": "Validation fail rõ, không side effect/partial write và không lộ stack trace",
            "recovery": "Sửa dữ liệu theo validation message; chạy lại idempotent; tạo defect nếu contract không rõ",
        },
        {
            "title": f"{subject} - dependency hoặc xử lý nội bộ thất bại",
            "trigger": "Dependency timeout/5xx hoặc exception nội bộ có kiểm soát",
            "steps": "Mock dependency lỗi; thực hiện lại scenario; kiểm tra response, log và trạng thái dữ liệu",
            "expected": f"Fail safe, có correlation ID và rollback; happy expectation không bị ghi sai: {expected}",
            "recovery": "Khôi phục dependency; retry theo policy; đối chiếu audit/evidence trước đóng defect",
        },
    ]


def _append_table(sheet, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    if rows:
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(headers))}{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 32


def _set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _add_validations(workbook: Workbook) -> None:
    master = workbook["01_Test_Case_Master"]
    priority = DataValidation(type="list", formula1='"P0,P1,P2,P3"')
    design = DataValidation(type="list", formula1='"Draft,Ready,Approved,Retired"')
    automation = DataValidation(type="list", formula1='"Automated,Partial,Planned,Manual"')
    for validation, column in ((priority, "I"), (design, "J"), (automation, "K")):
        master.add_data_validation(validation)
        validation.add(f"{column}2:{column}{master.max_row}")

    evidence = workbook["05_Execution_Evidence"]
    execution = DataValidation(type="list", formula1='"Passed,Failed,Blocked,Not Run"')
    evidence.add_data_validation(execution)
    execution.add(f"I2:I{max(2, evidence.max_row)}")


def build_workbook(source: Path, target: Path) -> dict[str, int]:
    cases = _load_legacy_cases(source)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    guide = workbook.create_sheet("00_Huong_Dan")
    guide.sheet_view.showGridLines = False
    guide["A1"] = "AVM PRODUCTION TEST CATALOGUE & RELEASE GATE"
    guide["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    guide["A1"].fill = HEADER_FILL
    guide.merge_cells("A1:F1")
    guide["A3"] = "Mục đích"
    guide["B3"] = "Tách test design khỏi test execution; mỗi scenario có ID, expected result, evidence và hướng xử lý riêng."
    guide["A5"] = "Sheet"
    guide["B5"] = "Vai trò"
    guide_rows = [
        ("01_Test_Case_Master", "Một dòng cho một test objective/requirement, ID duy nhất."),
        ("02_Happy_Path", "Kịch bản đầu ra mong muốn; có cách xử lý khi kết quả không đạt."),
        ("03_Failure_Path", "Hai hoặc nhiều negative/error scenario cho mỗi test case; không dồn chung một hàng."),
        ("04_Automation_Mapping", "Liên kết Test Case ID với command, node/path và CI job thực thi."),
        ("05_Execution_Evidence", "Kết quả chạy theo build/commit SHA; Actual Result tách khỏi design status."),
        ("06_Release_Gate", "Dashboard release readiness; P0 failed/blocked/not-run luôn là blocker."),
        ("07_RTM", "Traceability Requirement -> Risk -> Test Case -> Scenario -> Evidence."),
        ("08_Audit_Closure", "Bằng chứng đã đóng lỗi audit workbook v2."),
    ]
    for index, (name, purpose) in enumerate(guide_rows, start=6):
        guide.cell(index, 1, name)
        guide.cell(index, 2, purpose)
    guide["A16"] = "Quy tắc trạng thái"
    guide["B16"] = "Design Status: Draft/Ready/Approved/Retired. Execution Status: Passed/Failed/Blocked/Not Run. Không dùng Automated để thay cho Passed."
    guide["A18"] = "Release policy"
    guide["B18"] = "Không release khi P0 Failed/Blocked/Not Run, migration lệch head, security gate fail, hoặc active model không khớp registry/artifact."
    guide["A20"] = "Cập nhật"
    guide["B20"] = datetime.now().astimezone().isoformat(timespec="seconds")
    for cell in ("A3", "A5", "B5", "A16", "A18", "A20"):
        guide[cell].font = Font(bold=True)
        guide[cell].fill = SUBHEADER_FILL
    _set_widths(guide, {"A": 29, "B": 105, "C": 4, "D": 18, "E": 18, "F": 18})
    for row in guide.iter_rows(min_row=1, max_row=guide.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    master_headers = [
        "STT", "Test Case ID", "Requirement ID", "Risk Category", "Cấp kiểm thử",
        "Loại kiểm thử", "Khu vực", "Mục tiêu kiểm thử", "Priority", "Design Status",
        "Automation Status", "Owner", "Reviewer", "Environment", "Automation/Reference",
        "Evidence Policy",
    ]
    master_rows: list[list[Any]] = []
    happy_rows: list[list[Any]] = []
    failure_rows: list[list[Any]] = []
    automation_rows: list[list[Any]] = []
    evidence_rows: list[list[Any]] = []
    rtm_rows: list[list[Any]] = []

    for index, case in enumerate(cases, start=1):
        test_id = case["Mã test"]
        requirement = _requirement_id(test_id)
        risk = _risk_category(case)
        priority = _normalize_priority(case)
        automation_status, reference = _automation_status(case)
        if priority == "P0" and automation_status == "Planned":
            automation_status = "Manual"
        environment = "CI + PostgreSQL" if automation_status != "Planned" else "Staging production-like"
        owner = "QA Automation" if automation_status == "Automated" else "QA/Module Owner"
        evidence_policy = "JUnit/log + commit SHA" if automation_status == "Automated" else "Screenshot/log + reviewer sign-off"
        master_rows.append([
            index, test_id, requirement, risk, case["Cấp kiểm thử"], case["Loại kiểm thử"],
            case["Khu vực"], case["Kịch bản"], priority, "Approved", automation_status,
            owner, "Release Reviewer", environment, reference, evidence_policy,
        ])

        happy_id = f"{test_id}-H01"
        happy_rows.append([
            index, happy_id, test_id, "Happy path", case["Kịch bản"], "Dữ liệu hợp lệ, có nguồn và ID cố định",
            case["Điều kiện trước"], case["Bước thực hiện"], case["Kết quả mong đợi"],
            "Không có error ngoài contract; không có side effect thừa",
            "Mở defect kèm evidence; rollback dữ liệu nếu có write; sửa và rerun cùng fixture/build",
            priority, owner,
        ])

        failures = _failure_templates(case)
        failure_ids: list[str] = []
        for failure_index, failure in enumerate(failures, start=1):
            failure_id = f"{test_id}-F{failure_index:02d}"
            failure_ids.append(failure_id)
            failure_rows.append([
                len(failure_rows) + 1, failure_id, test_id, "Failure path", failure["title"],
                failure["trigger"], "Dùng fixture riêng; không dùng secret hoặc dữ liệu production thật",
                failure["steps"], failure["expected"], failure["recovery"],
                "Fail closed / rollback / evidence", priority, owner,
            ])

        command = ""
        ci_job = "manual-staging"
        if automation_status == "Automated":
            command = "python -m pytest tests -q" if "tests/" in reference else "python scripts/quality/validate_test_catalog.py"
            ci_job = "backend-tests" if "tests/" in reference else "test-catalogue"
        elif automation_status == "Partial":
            command = "Theo automation/reference; cần bổ sung assertion/evidence còn thiếu"
            ci_job = "mapped-partial"
        automation_rows.append([
            index, test_id, automation_status, command, reference, ci_job, environment,
            "JUnit XML + console log + artifact", "0 lần retry cho assertion; tối đa 1 lần cho lỗi hạ tầng đã xác nhận",
            owner,
        ])

        evidence_rows.append([
            index, test_id, happy_id, priority, environment, "", "", "", "Not Run", "", "", "",
            owner, "Release Reviewer",
        ])
        rtm_rows.append([
            index, requirement, risk, test_id, happy_id, ", ".join(failure_ids), priority,
            automation_status, reference, "05_Execution_Evidence",
        ])

    master = workbook.create_sheet("01_Test_Case_Master")
    _append_table(master, master_headers, master_rows, "TestCaseMaster")
    _set_widths(master, {
        "A": 7, "B": 22, "C": 20, "D": 24, "E": 17, "F": 18, "G": 24,
        "H": 54, "I": 10, "J": 16, "K": 18, "L": 18, "M": 18, "N": 22,
        "O": 54, "P": 36,
    })
    master.conditional_formatting.add(f"I2:I{master.max_row}", FormulaRule(formula=["I2=\"P0\""], fill=DANGER_FILL))
    master.conditional_formatting.add(f"K2:K{master.max_row}", FormulaRule(formula=["K2=\"Planned\""], fill=WARNING_FILL))

    happy = workbook.create_sheet("02_Happy_Path")
    _append_table(happy, [
        "STT", "Scenario ID", "Test Case ID", "Path Type", "Kịch bản", "Test Data / Fixture",
        "Điều kiện trước", "Bước thực hiện", "Kết quả mong đợi", "Điều không được xảy ra",
        "Hướng xử lý khi không đạt", "Priority", "Owner",
    ], happy_rows, "HappyPathScenarios")
    _set_widths(happy, {"A": 7, "B": 28, "C": 22, "D": 14, "E": 48, "F": 34, "G": 42, "H": 65, "I": 65, "J": 45, "K": 58, "L": 10, "M": 18})

    failure = workbook.create_sheet("03_Failure_Path")
    _append_table(failure, [
        "STT", "Scenario ID", "Test Case ID", "Path Type", "Kịch bản lỗi", "Điều kiện kích hoạt",
        "Test Data / Fixture", "Bước tái hiện", "Kết quả mong đợi", "Hướng xử lý / Recovery",
        "Failure Contract", "Priority", "Owner",
    ], failure_rows, "FailurePathScenarios")
    _set_widths(failure, {"A": 7, "B": 28, "C": 22, "D": 14, "E": 52, "F": 48, "G": 40, "H": 65, "I": 65, "J": 65, "K": 35, "L": 10, "M": 18})

    mapping = workbook.create_sheet("04_Automation_Mapping")
    _append_table(mapping, [
        "STT", "Test Case ID", "Automation Status", "Local Command", "Test Node / Reference",
        "GitHub Actions Job", "Environment", "Evidence Output", "Retry Policy", "Owner",
    ], automation_rows, "AutomationMapping")
    _set_widths(mapping, {"A": 7, "B": 22, "C": 18, "D": 45, "E": 62, "F": 24, "G": 24, "H": 42, "I": 48, "J": 20})

    evidence = workbook.create_sheet("05_Execution_Evidence")
    _append_table(evidence, [
        "STT", "Test Case ID", "Scenario ID", "Priority", "Environment", "Build / Commit SHA",
        "Last Run Date", "Actual Result", "Execution Status", "Evidence URL / Artifact",
        "Defect ID", "Ghi chú", "Owner", "Reviewer",
    ], evidence_rows, "ExecutionEvidence")
    _set_widths(evidence, {"A": 7, "B": 22, "C": 28, "D": 10, "E": 24, "F": 25, "G": 22, "H": 58, "I": 18, "J": 55, "K": 18, "L": 42, "M": 20, "N": 20})
    evidence.conditional_formatting.add(f"I2:I{evidence.max_row}", FormulaRule(formula=["I2=\"Passed\""], fill=ACCENT_FILL))
    evidence.conditional_formatting.add(f"I2:I{evidence.max_row}", FormulaRule(formula=["OR(I2=\"Failed\",I2=\"Blocked\")"], fill=DANGER_FILL))

    gate = workbook.create_sheet("06_Release_Gate")
    gate.sheet_view.showGridLines = False
    gate["A1"] = "AVM PRODUCTION RELEASE GATE"
    gate["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    gate["A1"].fill = HEADER_FILL
    gate.merge_cells("A1:E1")
    gate.append([])
    gate.append(["Gate", "Công thức / Nguồn", "Giá trị", "Ngưỡng", "Kết luận"])
    gate_rows = [
        ("Test Case ID unique", "01_Test_Case_Master", f"=COUNTA('01_Test_Case_Master'!B2:B{len(cases)+1})-SUMPRODUCT(1/COUNTIF('01_Test_Case_Master'!B2:B{len(cases)+1},'01_Test_Case_Master'!B2:B{len(cases)+1}))", "0", '=IF(C4=0,"PASS","BLOCK")'),
        ("P0 Failed", "05_Execution_Evidence", '=COUNTIFS(\'05_Execution_Evidence\'!D:D,"P0",\'05_Execution_Evidence\'!I:I,"Failed")', "0", '=IF(C5=0,"PASS","BLOCK")'),
        ("P0 Blocked", "05_Execution_Evidence", '=COUNTIFS(\'05_Execution_Evidence\'!D:D,"P0",\'05_Execution_Evidence\'!I:I,"Blocked")', "0", '=IF(C6=0,"PASS","BLOCK")'),
        ("P0 Not Run", "05_Execution_Evidence", '=COUNTIFS(\'05_Execution_Evidence\'!D:D,"P0",\'05_Execution_Evidence\'!I:I,"Not Run")', "0", '=IF(C7=0,"PASS","BLOCK")'),
        ("P0 Planned automation", "Master catalogue", '=COUNTIFS(\'01_Test_Case_Master\'!I:I,"P0",\'01_Test_Case_Master\'!K:K,"Planned")', "0 hoặc manual sign-off", '=IF(C8=0,"PASS","REVIEW")'),
        ("Failure scenario coverage", "03_Failure_Path", f"=COUNTA('03_Failure_Path'!B2:B{len(failure_rows)+1})", f">={len(cases)*2}", f'=IF(C9>={len(cases)*2},"PASS","BLOCK")'),
        ("PostgreSQL/Alembic", "CI backend-tests", "Nhập từ CI evidence", "head + tests pass", "PENDING EVIDENCE"),
        ("Security/third-party", "CI security + SonarCloud", "Nhập từ CI evidence", "all blocking gates pass", "PENDING EVIDENCE"),
        ("Docker smoke", "CI docker-smoke", "Nhập từ CI evidence", "all services healthy", "PENDING EVIDENCE"),
    ]
    for row in gate_rows:
        gate.append(row)
    gate["A14"] = "FINAL VERDICT"
    gate["C14"] = '=IF(COUNTIF(E4:E12,"BLOCK")>0,"NOT READY",IF(COUNTIF(E4:E12,"PENDING EVIDENCE")>0,"PENDING EVIDENCE","READY"))'
    gate["A14"].font = Font(bold=True)
    gate["C14"].font = Font(size=14, bold=True)
    for cell in gate[3]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in gate.iter_rows(min_row=4, max_row=14, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
    gate.conditional_formatting.add("E4:E14", FormulaRule(formula=['E4="PASS"'], fill=ACCENT_FILL))
    gate.conditional_formatting.add("E4:E14", FormulaRule(formula=['OR(E4="BLOCK",E4="NOT READY")'], fill=DANGER_FILL))
    _set_widths(gate, {"A": 32, "B": 36, "C": 30, "D": 30, "E": 22})

    rtm = workbook.create_sheet("07_RTM")
    _append_table(rtm, [
        "STT", "Requirement ID", "Risk Category", "Test Case ID", "Happy Scenario ID",
        "Failure Scenario IDs", "Priority", "Automation Status", "Test/Evidence Reference", "Execution Sheet",
    ], rtm_rows, "RequirementsTraceability")
    _set_widths(rtm, {"A": 7, "B": 22, "C": 25, "D": 22, "E": 28, "F": 58, "G": 10, "H": 18, "I": 65, "J": 24})

    closure = workbook.create_sheet("08_Audit_Closure")
    closure_rows = [
        [1, "Critical", "Rows 101, 102, 121 lệch cột", "Closed", "Chuẩn hóa ML-LIN-017, ML-LIN-018, CLEAN-004 bằng mapping tường minh", "scripts/quality/build_production_test_workbook.py"],
        [2, "Critical", "11 nhóm Test Case ID trùng", "Closed", "Đổi sang ID theo module; validator bắt buộc unique", "01_Test_Case_Master"],
        [3, "High", "Status trộn design và execution", "Closed", "Tách Design Status, Automation Status và Execution Status", "01_Test_Case_Master; 05_Execution_Evidence"],
        [4, "High", "Happy/failure path dồn trong một hàng", "Closed", "Mỗi scenario có ID riêng; tối thiểu 1 happy + 2 failure/test case", "02_Happy_Path; 03_Failure_Path"],
        [5, "High", "Thiếu Requirement/Risk/Owner/Reviewer/Build/Evidence", "Closed", "Bổ sung Master, RTM, Evidence và Release Gate", "01_Test_Case_Master; 05_Execution_Evidence; 07_RTM"],
        [6, "High", "Chưa có CI contract cho workbook", "Closed", "Validator cấu trúc chạy trong GitHub Actions và xuất report", "scripts/quality/validate_test_catalog.py; .github/workflows/ci.yml"],
        [7, "High", "P0 chưa có evidence", "Open release blocker", "Release Gate giữ NOT READY cho đến khi P0 Passed có commit SHA/evidence", "06_Release_Gate"],
    ]
    _append_table(closure, ["STT", "Severity", "Issue audit v2", "Trạng thái", "Cách xử lý", "Evidence"], closure_rows, "AuditClosure")
    _set_widths(closure, {"A": 7, "B": 14, "C": 48, "D": 22, "E": 70, "F": 62})

    _add_validations(workbook)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return {"test_cases": len(cases), "happy_scenarios": len(happy_rows), "failure_scenarios": len(failure_rows)}


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, required=True, help="Workbook catalogue/audit đầu vào")
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET, help="Workbook production đầu ra")
    args = parser.parse_args()
    result = build_workbook(args.source.resolve(), args.target.resolve())
    print(
        "Workbook production đã tạo: "
        f"{result['test_cases']} test cases, {result['happy_scenarios']} happy scenarios, "
        f"{result['failure_scenarios']} failure scenarios -> {args.target.resolve()}"
    )


if __name__ == "__main__":
    main()
