#!/usr/bin/env python3
"""Export CI evidence matrix from the production test workbook."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = PROJECT_ROOT / "docs" / "testing" / "AVM_Production_Test_Cases.xlsx"
DEFAULT_WORKFLOW = PROJECT_ROOT / ".github" / "workflows" / "ci.yml"
DEFAULT_REPORT = PROJECT_ROOT / "reports" / "ci" / "test-catalogue-ci-matrix.json"
DEFAULT_MARKDOWN = PROJECT_ROOT / "reports" / "ci" / "test-catalogue-ci-matrix.md"
AUTOMATION_SHEET = "04_Automation_Mapping"


def _records(sheet) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    return [
        dict(zip(headers, values, strict=False))
        for values in rows
        if any(value not in (None, "") for value in values)
    ]


def _normalize(value: Any) -> str:
    return str(value or "").strip()


def _workflow_job_keys(workflow_text: str) -> set[str]:
    match = re.search(r"(?ms)^jobs:\n(?P<body>.*?)(?:\n[a-zA-Z_].*:|\Z)", workflow_text)
    if not match:
        return set()
    return {
        item.group(1)
        for item in re.finditer(r"(?m)^  ([A-Za-z0-9_-]+):\s*$", match.group("body"))
    }


def _command_present(workflow_text: str, command: str) -> bool:
    normalized = " ".join(command.split())
    if normalized in " ".join(workflow_text.split()):
        return True
    if normalized.startswith("python scripts/"):
        return normalized.replace("python ", "python -m ") in workflow_text
    return False


def export_matrix(workbook_path: Path, workflow_path: Path) -> dict[str, Any]:
    workflow_text = workflow_path.read_text(encoding="utf-8")
    workflow_jobs = _workflow_job_keys(workflow_text)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if AUTOMATION_SHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Thiếu sheet {AUTOMATION_SHEET}")

    rows = _records(workbook[AUTOMATION_SHEET])
    workbook.close()

    errors: list[str] = []
    by_job: dict[str, list[dict[str, str]]] = defaultdict(list)
    by_status = Counter(_normalize(row.get("Automation Status")) for row in rows)
    command_to_ids: dict[str, list[str]] = defaultdict(list)

    for row in rows:
        test_id = _normalize(row.get("Test Case ID"))
        status = _normalize(row.get("Automation Status"))
        command = _normalize(row.get("Local Command"))
        reference = _normalize(row.get("Test Node / Reference"))
        job = _normalize(row.get("GitHub Actions Job"))

        if not test_id:
            errors.append("Mapping có dòng thiếu Test Case ID")
            continue
        if status == "Automated":
            if not command:
                errors.append(f"{test_id}: Automated nhưng thiếu Local Command")
            if not job or job == "mapped-partial":
                errors.append(f"{test_id}: Automated nhưng thiếu GitHub Actions Job thật")
            if job and job not in workflow_jobs:
                errors.append(f"{test_id}: job '{job}' không tồn tại trong workflow")
            if command and not _command_present(workflow_text, command):
                errors.append(f"{test_id}: command '{command}' chưa xuất hiện trong workflow")
            command_to_ids[command].append(test_id)

        by_job[job or "unmapped"].append(
            {
                "test_case_id": test_id,
                "automation_status": status,
                "local_command": command,
                "test_node_reference": reference,
                "evidence_output": _normalize(row.get("Evidence Output")),
                "environment": _normalize(row.get("Environment")),
            }
        )

    return {
        "valid": not errors,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workbook": str(workbook_path),
        "workflow": str(workflow_path),
        "errors": errors,
        "metrics": {
            "total_mapping_rows": len(rows),
            "automated": by_status.get("Automated", 0),
            "partial": by_status.get("Partial", 0),
            "planned": by_status.get("Planned", 0),
            "manual": by_status.get("Manual", 0),
            "workflow_jobs": sorted(workflow_jobs),
            "automated_command_count": len([cmd for cmd in command_to_ids if cmd]),
        },
        "automated_commands": [
            {"local_command": command, "test_case_ids": sorted(test_ids)}
            for command, test_ids in sorted(command_to_ids.items())
            if command
        ],
        "jobs": {job: rows for job, rows in sorted(by_job.items())},
    }


def _short(value: str, limit: int = 96) -> str:
    value = value.replace("|", "\\|").replace("\n", " ").strip()
    return value if len(value) <= limit else f"{value[:limit - 3]}..."


def render_markdown(result: dict[str, Any], job_filter: str | None = None) -> str:
    """Render a GitHub Actions summary that exposes workbook test IDs directly."""
    metrics = result["metrics"]
    title_suffix = f" — `{job_filter}`" if job_filter else ""
    lines = [
        f"## Production Test Catalogue Evidence{title_suffix}",
        "",
        "| Metric | Value |",
        "| --- | ---: |",
        f"| Tổng mapping rows | {metrics['total_mapping_rows']} |",
        f"| Automated | {metrics['automated']} |",
        f"| Partial | {metrics['partial']} |",
        f"| Planned | {metrics['planned']} |",
        f"| Manual | {metrics['manual']} |",
        f"| Workflow jobs được kiểm tra | {len(metrics['workflow_jobs'])} |",
        "",
        "### Automated Test Cases chạy trong CI",
        "",
    ]

    automated_rows: list[tuple[str, dict[str, str]]] = []
    for job, rows in result["jobs"].items():
        if job_filter and job != job_filter:
            continue
        for row in rows:
            if row["automation_status"] == "Automated":
                automated_rows.append((job, row))

    if not automated_rows:
        lines.append("_Không có testcase Automated trong workbook._")
    else:
        lines.extend([
            "| Test Case ID | GitHub job | Local command | Evidence |",
            "| --- | --- | --- | --- |",
        ])
        for job, row in sorted(automated_rows, key=lambda item: (item[0], item[1]["test_case_id"])):
            lines.append(
                "| "
                f"{_short(row['test_case_id'], 32)} | "
                f"{_short(job, 36)} | "
                f"`{_short(row['local_command'], 72)}` | "
                f"{_short(row['evidence_output'], 72)} |"
            )

    if result["errors"]:
        lines.extend(["", "### Blocking Errors", ""])
        lines.extend(f"- {error}" for error in result["errors"])

    lines.extend([
        "",
        "> CI này được sinh từ sheet `04_Automation_Mapping` trong workbook production testcase; "
        "nếu Test Case ID trùng, thiếu command, hoặc job không tồn tại thì job sẽ fail.",
        "",
    ])
    return "\n".join(lines)


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--workflow", type=Path, default=DEFAULT_WORKFLOW)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--markdown", type=Path, default=None)
    parser.add_argument("--job", type=str, default=None, help="Only render Markdown rows for one GitHub job")
    args = parser.parse_args()

    result = export_matrix(args.workbook.resolve(), args.workflow.resolve())
    output = json.dumps(result, ensure_ascii=False, indent=2)
    print(output)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(output + "\n", encoding="utf-8")
    if args.markdown:
        args.markdown.parent.mkdir(parents=True, exist_ok=True)
        args.markdown.write_text(render_markdown(result, job_filter=args.job), encoding="utf-8")
    raise SystemExit(0 if result["valid"] else 1)


if __name__ == "__main__":
    main()
