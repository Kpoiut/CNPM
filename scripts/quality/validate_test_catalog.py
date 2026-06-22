#!/usr/bin/env python3
"""Validate cấu trúc và traceability của AVM production test workbook."""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = PROJECT_ROOT / "docs" / "testing" / "AVM_Production_Test_Cases.xlsx"
REQUIRED_SHEETS = {
    "00_Huong_Dan",
    "01_Test_Case_Master",
    "02_Happy_Path",
    "03_Failure_Path",
    "04_Automation_Mapping",
    "05_Execution_Evidence",
    "06_Release_Gate",
    "07_RTM",
    "08_Audit_Closure",
}


def _records(sheet) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    return [dict(zip(headers, values, strict=False)) for values in rows if any(value not in (None, "") for value in values)]


def validate(workbook_path: Path) -> dict[str, Any]:
    errors: list[str] = []
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    missing_sheets = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing_sheets:
        errors.append(f"Thiếu sheet bắt buộc: {missing_sheets}")
        workbook.close()
        return {"valid": False, "errors": errors}

    master = _records(workbook["01_Test_Case_Master"])
    happy = _records(workbook["02_Happy_Path"])
    failure = _records(workbook["03_Failure_Path"])
    mapping = _records(workbook["04_Automation_Mapping"])
    evidence = _records(workbook["05_Execution_Evidence"])
    rtm = _records(workbook["07_RTM"])

    if len(master) < 100:
        errors.append(f"Catalogue phải có ít nhất 100 test cases, hiện có {len(master)}")
    stt = [row.get("STT") for row in master]
    if stt != list(range(1, len(master) + 1)):
        errors.append("STT master phải liên tục từ 1")

    test_ids = [str(row.get("Test Case ID") or "").strip() for row in master]
    duplicate_ids = sorted(test_id for test_id, count in Counter(test_ids).items() if count > 1)
    if duplicate_ids:
        errors.append(f"Test Case ID bị trùng: {duplicate_ids}")
    if any(not test_id for test_id in test_ids):
        errors.append("Test Case ID không được rỗng")
    master_ids = set(test_ids)

    for row in master:
        required = (
            "Requirement ID", "Risk Category", "Mục tiêu kiểm thử", "Priority", "Design Status",
            "Automation Status", "Owner", "Reviewer", "Environment", "Evidence Policy",
        )
        missing = [column for column in required if not row.get(column)]
        if missing:
            errors.append(f"{row.get('Test Case ID')}: thiếu {missing}")
        if row.get("Priority") == "P0" and row.get("Design Status") != "Approved":
            errors.append(f"{row.get('Test Case ID')}: P0 chưa Approved")

    happy_counts = Counter(str(row.get("Test Case ID")) for row in happy)
    failure_counts = Counter(str(row.get("Test Case ID")) for row in failure)
    scenario_ids = [str(row.get("Scenario ID") or "") for row in happy + failure]
    scenario_duplicates = sorted(value for value, count in Counter(scenario_ids).items() if count > 1)
    if scenario_duplicates:
        errors.append(f"Scenario ID bị trùng: {scenario_duplicates}")

    for test_id in test_ids:
        if happy_counts[test_id] < 1:
            errors.append(f"{test_id}: thiếu happy path")
        if failure_counts[test_id] < 2:
            errors.append(f"{test_id}: cần tối thiểu 2 failure paths")

    failure_required = ("Kịch bản lỗi", "Điều kiện kích hoạt", "Bước tái hiện", "Kết quả mong đợi", "Hướng xử lý / Recovery")
    for row in failure:
        missing = [column for column in failure_required if not row.get(column)]
        if missing:
            errors.append(f"{row.get('Scenario ID')}: thiếu {missing}")

    for sheet_name, rows in (("04_Automation_Mapping", mapping), ("05_Execution_Evidence", evidence), ("07_RTM", rtm)):
        row_ids = {str(row.get("Test Case ID") or "") for row in rows}
        missing = sorted(master_ids - row_ids)
        unknown = sorted(row_ids - master_ids)
        if missing:
            errors.append(f"{sheet_name}: thiếu mapping cho {missing}")
        if unknown:
            errors.append(f"{sheet_name}: chứa Test Case ID không tồn tại {unknown}")

    for row in mapping:
        if row.get("Automation Status") == "Automated" and not row.get("Local Command"):
            errors.append(f"{row.get('Test Case ID')}: Automated nhưng thiếu Local Command")
        if row.get("Automation Status") == "Automated" and not row.get("Test Node / Reference"):
            errors.append(f"{row.get('Test Case ID')}: Automated nhưng thiếu Test Node / Reference")

    workbook.close()
    return {
        "valid": not errors,
        "errors": errors,
        "metrics": {
            "test_cases": len(master),
            "happy_scenarios": len(happy),
            "failure_scenarios": len(failure),
            "p0": sum(1 for row in master if row.get("Priority") == "P0"),
            "automated": sum(1 for row in master if row.get("Automation Status") == "Automated"),
            "partial": sum(1 for row in master if row.get("Automation Status") == "Partial"),
            "planned": sum(1 for row in master if row.get("Automation Status") == "Planned"),
            "p0_planned": sum(
                1
                for row in master
                if row.get("Priority") == "P0" and row.get("Automation Status") == "Planned"
            ),
            "p0_manual": sum(
                1
                for row in master
                if row.get("Priority") == "P0" and row.get("Automation Status") == "Manual"
            ),
        },
    }


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", nargs="?", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--report", type=Path, help="Ghi JSON evidence cho CI artifact")
    args = parser.parse_args()
    result = validate(args.workbook.resolve())
    output = json.dumps(result, ensure_ascii=False, indent=2)
    print(output)
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(output + "\n", encoding="utf-8")
    raise SystemExit(0 if result["valid"] else 1)


if __name__ == "__main__":
    main()
